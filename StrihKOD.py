#Библиотеки
import os
import cv2
import base64
from pyzbar import pyzbar
from barcode import Code128
from barcode.writer import ImageWriter
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageFont, ImageDraw

# 1. Генерация штрих-кода ======================================================
def generate_barcode(phone, document, filename):
    # Кодируем русский текст в Base64
    encoded_doc = base64.b64encode(document.encode('utf-8')).decode('utf-8')
    
    # Формируем данные для штрих-кода
    barcode_data = f"{phone}${encoded_doc}"  # Используем $ как разделитель
    
    # Создаем и сохраняем штрих-код
    code = Code128(barcode_data, writer=ImageWriter())
    code.save(filename, options={"write_text": False})
    
    # Добавляем текст под штрих-кодом (опционально)
    add_text_to_barcode(f"{filename}.png", phone, document)

def add_text_to_barcode(image_path, phone, document):
    img = Image.open(image_path)
    draw = ImageDraw.Draw(img)
    
    try:
        font = ImageFont.truetype("arial.ttf", 16)
    except:
        font = ImageFont.load_default()
    
    text = f"Тел: {phone}\nДок: {document}"
    text_width, text_height = draw.textsize(text, font=font)
    
    # Позиция текста по центру внизу
    x = (img.width - text_width) / 2
    y = img.height - text_height - 10
    
    draw.text((x, y), text, font=font, fill="black")
    img.save(image_path)

# 2. Сканирование штрих-кода =================================================
def scan_barcode():
    cap = cv2.VideoCapture(0)
    
    while True:
        ret, frame = cap.read()
        barcodes = pyzbar.decode(frame)
        
        if barcodes:
            for barcode in barcodes:
                data = barcode.data.decode("utf-8")
                cap.release()
                cv2.destroyAllWindows()
                return data
        
        cv2.imshow("Scan Barcode (Q to quit)", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    
    cap.release()
    cv2.destroyAllWindows()
    return None

# 3. Работа с Excel ===========================================================
def update_excel(filename, phone, document):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Дата", "Номер телефона", "Документ"])
    
    from datetime import datetime
    date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    ws.append([date, phone, document])
    wb.save(filename)

# Основная программа ==========================================================
if __name__ == "__main__":
    # Генерация штрих-кода
    phone_number = "+79123456789"
    document_name = "Договор оказания услуг"
    generate_barcode(phone_number, document_name, "barcode_temp")
    
    # Сканирование
    print("Наведите камеру на штрих-код...")
    scanned_data = scan_barcode()
    
    if scanned_data:
        try:
            # Разделяем и декодируем данные
            parts = scanned_data.split("$")
            if len(parts) != 2:
                raise ValueError("Неправильный формат данных")
            
            decoded_phone = parts[0]
            decoded_document = base64.b64decode(parts[1]).decode('utf-8')
            
            # Запись в Excel
            update_excel("barcode_data.xlsx", decoded_phone, decoded_document)
            print("Данные успешно записаны в Excel!")
            
        except Exception as e:
            print(f"Ошибка обработки данных: {str(e)}")
    else:
        print("Штрих-код не отсканирован")

    # Удаление временного файла
    if os.path.exists("barcode_temp.png"):
        os.remove("barcode_temp.png")